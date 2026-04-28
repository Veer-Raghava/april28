"""
chatbot.py — Conversational chatbot for the Dataset Builder.

Key improvements:
  - Asks user for domain context at start (used to sharpen extraction prompts)
  - PDF-only source counting (HTML sources are bonus, not counted)
  - Live row display as each row is extracted
  - Post-run interactive analysis: stats, pivot, export options
  - Blocked sources summary with reasons
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

from tools.console_setup import console
from rich.panel import Panel
from rich.prompt import Prompt, Confirm
from rich.text import Text
from rich.markdown import Markdown
from rich.table import Table
from rich import box, print as rprint

import config as cfg
from tools.llm_client import LLMClient
from agents.orchestrator import run_pipeline_from_config, PipelineResult


MEMORY_FILE = os.path.join(cfg.MEMORY_DIR, "conversation_history.json")

BANNER = """[bold bright_cyan]
    ╔══════════════════════════════════════════════════════╗
    ║                                                      ║
    ║   🧬  Dataset Builder — AI Research Assistant        ║
    ║                                                      ║
    ║   Search · Scrape · Extract · Validate · Export      ║
    ║                                                      ║
    ╚══════════════════════════════════════════════════════╝
[/bold bright_cyan]
[dim]  Powered by 6 specialized AI agents
  Table-first extraction · Science API null hunting
  Anti-bot stealth · Live row append · ADC-domain aware
  Type /help for commands[/dim]
"""

HELP_TEXT = """
[bold cyan]Available Commands:[/bold cyan]

  [green]/help[/green]       — Show this help message
  [green]/status[/green]     — Show current pipeline status
  [green]/columns[/green]    — Show/modify current columns
  [green]/memory[/green]     — Show conversation memory
  [green]/export[/green]     — Re-export current data
  [green]/blocked[/green]    — Show all blocked/rejected sources
  [green]/stats[/green]      — Show dataset statistics
  [green]/reset[/green]      — Reset current session
  [green]/quit[/green]       — Exit the chatbot

[bold cyan]Natural language:[/bold cyan]

  "Build a dataset about antibody drug conjugates"
  "Add a column for DAR value"
  "Scrape 20 sources"
  "Show me what we have so far"
  "What are the blocked sources?"
  "Give me stats on the data"
"""


class MemoryManager:
    def __init__(self):
        self.history: list[dict] = []
        self._load()

    def _load(self):
        if os.path.exists(MEMORY_FILE):
            try:
                data = json.loads(Path(MEMORY_FILE).read_text(encoding="utf-8"))
                self.history = data.get("history", [])
            except Exception:
                self.history = []

    def save(self):
        Path(MEMORY_FILE).parent.mkdir(parents=True, exist_ok=True)
        data = {"history": self.history[-100:]}
        Path(MEMORY_FILE).write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")

    def add(self, role: str, content: str, metadata: dict | None = None):
        self.history.append({
            "role": role, "content": content,
            "timestamp": time.time(), "metadata": metadata or {},
        })
        self.save()

    def get_context(self, n: int = 10) -> str:
        return "\n".join(
            f"[{e['role'].upper()}]: {e['content'][:200]}"
            for e in self.history[-n:]
        )

    def get_past_topics(self) -> list[str]:
        return list({e["metadata"]["topic"] for e in self.history
                     if e.get("metadata", {}).get("topic")})

    def get_past_columns(self, topic: str) -> list[str]:
        for e in reversed(self.history):
            if e.get("metadata", {}).get("columns") and topic.lower() in e.get("content", "").lower():
                return e["metadata"]["columns"]
        return []


def parse_columns_from_file(filepath: str) -> list[str]:
    path = Path(filepath)
    if not path.exists():
        console.print(f"[red]✗ File not found: {filepath}[/red]")
        return []
    content = path.read_text(encoding="utf-8").strip()
    ext = path.suffix.lower()
    if ext == ".json":
        try:
            data = json.loads(content)
            if isinstance(data, list):
                return ([d.get("name", "") for d in data if isinstance(d, dict)]
                        if data and isinstance(data[0], dict)
                        else [str(c) for c in data if c])
            if isinstance(data, dict):
                for key in ["columns", "fields", "names"]:
                    if key in data and isinstance(data[key], list):
                        items = data[key]
                        return ([d.get("name", "") for d in items if isinstance(d, dict)]
                                if items and isinstance(items[0], dict)
                                else [str(c) for c in items if c])
        except json.JSONDecodeError:
            return []
    elif ext == ".csv":
        lines = content.split("\n")
        if lines:
            return [c.strip().strip('"').strip("'") for c in lines[0].split(",") if c.strip()]
    else:
        return [l.strip() for l in content.split("\n") if l.strip() and not l.startswith("#")]
    return []


class DatasetChatbot:

    def __init__(self):
        self.client: LLMClient | None = None
        self.memory = MemoryManager()
        self.state: PipelineResult | None = None
        self.topic: str = ""
        self.domain_context: str = ""   # NEW: user's domain knowledge injected into prompts
        self.columns: list[str] = []
        self.max_sources: int = cfg.DEFAULT_LIMIT  # number of PDFs (HTML is extra)
        self.running: bool = False

    def start(self):
        console.print(BANNER)
        try:
            self.client = LLMClient()
        except (ValueError, SystemExit) as e:
            console.print(f"[red]✗ {e}[/red]")
            sys.exit(1)
        self.memory.add("system", "Chatbot session started")
        past_topics = self.memory.get_past_topics()
        if past_topics:
            console.print(f"\n  [dim]📚 Past topics: {', '.join(past_topics[-5:])}[/dim]")
        self._greet()
        self._conversation_loop()

    def _greet(self):
        console.print()
        console.print(Panel(
            "[bold bright_white]Hey! 👋 I'm your AI Dataset Builder.[/bold bright_white]\n\n"
            "Tell me what kind of dataset you want to build and I'll handle the rest.\n"
            "I'll search the web, download papers, extract structured rows, and export a clean CSV.\n\n"
            "[dim]Try: \"Build a dataset about antibody drug conjugates\"[/dim]",
            border_style="bright_cyan",
            padding=(1, 2),
        ))

    def _conversation_loop(self):
        while True:
            try:
                console.print()
                user_input = Prompt.ask("[bold cyan]You[/bold cyan]").strip()
                if not user_input:
                    continue
                self.memory.add("user", user_input)
                if user_input.startswith("/"):
                    if self._handle_command(user_input) == "quit":
                        break
                    continue
                self._handle_natural_input(user_input)
            except KeyboardInterrupt:
                console.print("\n[dim]Use /quit to exit[/dim]")
            except EOFError:
                break
        self._goodbye()

    def _handle_command(self, cmd: str) -> str:
        c = cmd.lower().strip()
        if c in ("/quit", "/exit", "/q"):      return "quit"
        elif c in ("/help", "/h", "/?"):       console.print(HELP_TEXT)
        elif c in ("/status", "/s"):           self._show_status()
        elif c in ("/columns", "/cols"):       self._show_columns()
        elif c in ("/memory", "/mem"):         self._show_memory()
        elif c in ("/export",):                self._export()
        elif c in ("/blocked",):               self._show_blocked()
        elif c in ("/stats",):                 self._show_stats()
        elif c in ("/reset",):                 self._reset()
        else:
            console.print(f"[yellow]Unknown command: {cmd}. Type /help.[/yellow]")
        return ""

    def _handle_natural_input(self, message: str):
        ml = message.lower()
        if not self.topic:
            self._set_topic(message)
        elif not self.columns:
            self._handle_column_input(message)
        elif not self.running and any(kw in ml for kw in ["start", "go", "run", "build", "scrape", "begin"]):
            self._run_pipeline()
        elif "add column" in ml:
            col = re.sub(r"(?:add\s+(?:a\s+)?column\s+(?:for\s+|named\s+)?)", "", message, flags=re.I).strip()
            if col:
                self.columns.append(col)
                console.print(f"\n  [green]✓ Added column: [bold]{col}[/bold][/green]")
                self._show_columns()
        elif any(kw in ml for kw in ["how many source", "sources", "scrape more"]):
            nums = re.findall(r"\d+", message)
            if nums:
                self.max_sources = int(nums[0])
                console.print(f"\n  [green]✓ PDF source target: [bold]{self.max_sources}[/bold][/green]")
                console.print("  [dim](HTML sources are collected additionally)[/dim]")
        elif any(kw in ml for kw in ["blocked", "rejected", "paywall"]):
            self._show_blocked()
        elif any(kw in ml for kw in ["stat", "overview", "summary", "how good"]):
            self._show_stats()
        elif any(kw in ml for kw in ["show", "preview", "what we have", "results"]):
            self._show_status()
        else:
            self._smart_respond(message)

    def _set_topic(self, message: str):
        topic = message
        for prefix in ["build a dataset about ", "create a dataset for ", "i want data on ",
                       "find data about ", "research ", "dataset for ", "dataset about "]:
            if topic.lower().startswith(prefix):
                topic = topic[len(prefix):]
                break
        self.topic = topic.strip().strip('"').strip("'")
        self.memory.add("assistant", f"Topic set: {self.topic}", {"topic": self.topic})
        console.print(f"\n  [green]✓ Topic: [bold]{self.topic}[/bold][/green]")

        paths = cfg.init_topic_workspace(self.topic)
        console.print(f"  [dim]📂 Workspace: {paths['base_dir']}[/dim]")

        existing_csv = Path(paths["output_file"])
        if existing_csv.exists() and existing_csv.stat().st_size > 0:
            console.print(f"  [green]📚 Existing dataset found — new rows will be appended.[/green]")

        # ── NEW: Ask for domain context to sharpen extraction ────────────
        console.print()
        console.print(Panel(
            "[bold]Tell me about your domain context (optional but very helpful).[/bold]\n\n"
            "For example:\n"
            '  "I work in ADC biopharma. Key fields are DAR, MMAE, HER2, cleavable linkers."\n'
            '  "I study CRISPR gene editing. Important: guide RNA, PAM sequence, off-target."\n\n'
            "[dim]This helps me extract the right values and fill in nulls intelligently.[/dim]",
            border_style="cyan",
            padding=(1, 2),
        ))
        self.domain_context = Prompt.ask(
            "  [bold]Domain context[/bold]",
            default="(skip)"
        ).strip()
        if self.domain_context.lower() in ("(skip)", "skip", ""):
            self.domain_context = ""
        else:
            console.print(f"  [green]✓ Domain context saved[/green]")

        past_cols = self.memory.get_past_columns(self.topic)
        if past_cols:
            console.print(f"\n  [dim]📚 Found columns from a similar past session:[/dim]")
            for c in past_cols:
                console.print(f"    [dim]• {c}[/dim]")
            if Confirm.ask("  [bold]Use these columns?[/bold]", default=True):
                self.columns = past_cols
                self._show_columns()
                self._ask_ready()
                return
        self._ask_columns()

    def _ask_columns(self):
        console.print()
        console.print(Panel(
            "[bold]How would you like to define your columns?[/bold]\n\n"
            "  [cyan]1.[/cyan] Type them manually\n"
            "  [cyan]2.[/cyan] Load from a file (.txt, .csv, .json)\n"
            "  [cyan]3.[/cyan] Let AI suggest them",
            border_style="cyan", padding=(1, 2),
        ))
        choice = Prompt.ask("  [bold]Choose[/bold]", choices=["1", "2", "3"], default="1")
        if choice == "1":
            console.print("\n  [dim]Enter column names (comma-separated or one per line). Type 'done' to finish.[/dim]\n")
            self._manual_column_input()
        elif choice == "2":
            self._load_columns_from_file(Prompt.ask("  [bold]File path[/bold]"))
        elif choice == "3":
            self._ai_suggest_columns()

    def _manual_column_input(self):
        cols: list[str] = []
        while True:
            line = Prompt.ask(f"  [cyan]Column {len(cols)+1}[/cyan]", default="done")
            if line.lower() in ("done", "d", ""):
                break
            for c in line.split(","):
                c = c.strip().strip('"').strip("'")
                if c and c.lower() != "done":
                    cols.append(c)
                    console.print(f"    [green]+ {c}[/green]")
        if cols:
            self.columns = cols
            self._show_columns()
            self.memory.add("assistant", f"Columns set: {cols}", {"columns": cols})
            self._ask_ready()
        else:
            console.print("  [yellow]No columns entered. Try /help.[/yellow]")

    def _load_columns_from_file(self, filepath: str):
        cols = parse_columns_from_file(filepath.strip().strip('"').strip("'"))
        if cols:
            self.columns = cols
            console.print(f"\n  [green]✓ Loaded {len(cols)} columns from {filepath}[/green]")
            self._show_columns()
            self.memory.add("assistant", f"Columns loaded from {filepath}", {"columns": cols})
            self._ask_ready()
        else:
            console.print("  [yellow]Could not load columns. Check the file.[/yellow]")

    def _ai_suggest_columns(self):
        console.print("\n  [dim]🤔 Asking AI to suggest columns…[/dim]")
        ctx = f"\n\nDomain context: {self.domain_context}" if self.domain_context else ""
        system = (
            "Suggest 8-12 column names to extract from primary research papers about this topic. "
            "Focus on measurable, experimentally reported values. "
            "Return ONLY a JSON array of short column name strings."
        )
        result = self.client.complete_json(
            f'Topic: "{self.topic}"{ctx}\n\nSuggest columns:',
            system=system, max_tokens=512,
        )
        if isinstance(result, list) and result:
            suggested = [str(c) for c in result if isinstance(c, str)]
            if suggested:
                console.print(f"\n  [bold]AI suggested {len(suggested)} columns:[/bold]")
                for i, c in enumerate(suggested, 1):
                    console.print(f"    [cyan]{i}.[/cyan] {c}")
                console.print()
                if Confirm.ask("  [bold]Accept all?[/bold]", default=True):
                    self.columns = suggested
                else:
                    edits = Prompt.ask("  [bold]Edit (comma-separated list)[/bold]")
                    new_cols = [c.strip() for c in edits.split(",") if c.strip()]
                    self.columns = new_cols if new_cols else suggested
                self._show_columns()
                self.memory.add("assistant", f"AI columns: {self.columns}", {"columns": self.columns})
                self._ask_ready()
                return
        console.print("  [yellow]AI suggestion failed. Entering columns manually.[/yellow]")
        self._manual_column_input()

    def _handle_column_input(self, message: str):
        if message.lower() in ("done", "ok", "ready"):
            if self.columns:
                self._ask_ready()
            else:
                console.print("  [yellow]No columns yet! Type column names first.[/yellow]")
        else:
            for c in message.split(","):
                c = c.strip().strip('"')
                if c:
                    self.columns.append(c)
                    console.print(f"    [green]+ {c}[/green]")
            console.print(f"  [dim]({len(self.columns)} columns so far. Type 'done' to finish)[/dim]")

    def _ask_ready(self):
        console.print()
        count = Prompt.ask(
            "  [bold]How many PDFs to collect?[/bold]\n"
            "  [dim](HTML sources are collected in addition — this is PDF count only)[/dim]\n"
            "  [bold]>[/bold]",
            default=str(self.max_sources),
        )
        try:
            self.max_sources = int(count)
        except ValueError:
            pass

        console.print()
        console.print(Panel(
            f"[bold]Ready to build dataset![/bold]\n\n"
            f"  📌 Topic:   [cyan]{self.topic}[/cyan]\n"
            f"  📊 Columns: [cyan]{len(self.columns)}[/cyan] — "
            f"{', '.join(self.columns[:5])}{'…' if len(self.columns) > 5 else ''}\n"
            f"  📄 PDF target: [cyan]{self.max_sources}[/cyan] sources\n"
            f"  🔧 Provider: [cyan]{cfg.LLM_PROVIDER} / {cfg.active_model()}[/cyan]"
            + (f"\n  🧪 Domain context: [dim]{self.domain_context[:60]}[/dim]"
               if self.domain_context else ""),
            border_style="bright_green", padding=(1, 2),
        ))
        if Confirm.ask("  [bold]Start scraping?[/bold]", default=True):
            self._run_pipeline()
        else:
            console.print("  [dim]OK. Type /help for options.[/dim]")

    def _run_pipeline(self):
        self.running = True
        config = {
            "dataset_name":       self.topic,
            "description":        self.topic,
            "columns":            [{"name": c} for c in self.columns],
            "min_rows":           50,
            "max_sources":        self.max_sources,
            "max_adaptive_loops": cfg.MAX_ADAPTIVE_LOOPS,
            "domain_context":     self.domain_context,
        }
        try:
            self.state = run_pipeline_from_config(config)
            n_accepted = len(self.state.accepted_rows)
            n_blocked  = len(self.state.blocked_sources)
            self.memory.add(
                "assistant",
                f"Pipeline done for '{self.topic}': {n_accepted} rows, {n_blocked} blocked",
                {"topic": self.topic, "columns": self.columns, "rows": n_accepted},
            )
            console.print(f"\n  [bold green]🎉 Done! {n_accepted} rows extracted.[/bold green]")
            console.print(f"  [dim]Output: {self.state.output_path}[/dim]")
            if n_blocked > 0:
                console.print(f"  [yellow]⚠ {n_blocked} source(s) blocked (paywall/bot). Type /blocked to see them.[/yellow]")
            # Show post-run options
            self._post_run_options()
        except Exception as e:
            console.print(f"\n  [red]✗ Pipeline error: {e}[/red]")
        self.running = False

    def _post_run_options(self):
        """Interactive post-run menu."""
        console.print()
        console.print(Panel(
            "[bold]What would you like to do next?[/bold]\n\n"
            "  [cyan]1.[/cyan] Show data statistics\n"
            "  [cyan]2.[/cyan] Show blocked sources\n"
            "  [cyan]3.[/cyan] Run again for more rows (will append)\n"
            "  [cyan]4.[/cyan] Export to Excel\n"
            "  [cyan]5.[/cyan] Continue chatting\n"
            "  [cyan]6.[/cyan] Start a new topic",
            border_style="bright_cyan", padding=(1, 2),
        ))
        choice = Prompt.ask(
            "  [bold]Choose[/bold]",
            choices=["1", "2", "3", "4", "5", "6"], default="5"
        )
        if choice == "1":
            self._show_stats()
        elif choice == "2":
            self._show_blocked()
        elif choice == "3":
            extra = Prompt.ask("  [bold]How many more PDFs?[/bold]", default="10")
            try:
                self.max_sources = int(extra)
            except ValueError:
                pass
            self._run_pipeline()
        elif choice == "4":
            self._export_excel()
        elif choice == "6":
            self.topic = ""
            self.columns = []
            self.state = None
            console.print("  [green]✓ Ready for a new topic.[/green]")

    # ── Display helpers ───────────────────────────────────────────────────────

    def _show_status(self):
        if not self.state:
            console.print(Panel(
                f"  📌 Topic: [cyan]{self.topic or 'Not set'}[/cyan]\n"
                f"  📊 Columns: [cyan]{len(self.columns)}[/cyan]\n"
                f"  📄 PDF target: [cyan]{self.max_sources}[/cyan]\n"
                f"  🔧 Provider: [cyan]{cfg.LLM_PROVIDER} / {cfg.active_model()}[/cyan]",
                title="[bold]📋 Session Status[/bold]",
                border_style="bright_blue", padding=(1, 2),
            ))
        else:
            null_pct = f"{self.state.null_rate:.0%}" if self.state.null_rate is not None else "N/A"
            console.print(Panel(
                f"  📌 Topic: [cyan]{self.topic}[/cyan]\n"
                f"  [green]✓ Accepted:[/green] {len(self.state.accepted_rows)} rows\n"
                f"  [yellow]✗ Rejected:[/yellow] {len(self.state.rejected_rows)} rows\n"
                f"  [red]🛡 Blocked:[/red]  {len(self.state.blocked_sources)} sources\n"
                f"  📡 Documents: {len(self.state.documents)} processed\n"
                f"  📉 Null rate: {null_pct}\n"
                f"  💾 Output: {self.state.output_path}",
                title="[bold]📋 Pipeline Status[/bold]",
                border_style="bright_green", padding=(1, 2),
            ))

    def _show_columns(self):
        if not self.columns:
            console.print("  [dim]No columns defined yet.[/dim]")
            return
        t = Table(title=f"[bold cyan]📊 Columns ({len(self.columns)})[/bold cyan]",
                  show_lines=False, border_style="cyan", box=box.SIMPLE)
        t.add_column("#", style="dim", width=3)
        t.add_column("Column Name", style="bright_white")
        for i, col in enumerate(self.columns, 1):
            t.add_row(str(i), col)
        console.print(t)

    def _show_memory(self):
        if not self.memory.history:
            console.print("  [dim]No conversation history yet.[/dim]")
            return
        t = Table(title=f"[bold magenta]📚 Memory ({len(self.memory.history)} entries)[/bold magenta]",
                  show_lines=False, border_style="magenta", box=box.SIMPLE)
        t.add_column("Role", style="cyan", width=10)
        t.add_column("Content", max_width=60, overflow="ellipsis")
        for e in self.memory.history[-15:]:
            t.add_row(e["role"], e["content"][:60])
        console.print(t)

    def _show_blocked(self):
        """Show all blocked sources with reasons."""
        if not self.state or not self.state.blocked_sources:
            console.print("  [dim]No blocked sources recorded.[/dim]")
            return
        t = Table(
            title=f"[bold red]🛡 Blocked Sources ({len(self.state.blocked_sources)})[/bold red]",
            show_lines=True, border_style="red", box=box.ROUNDED,
        )
        t.add_column("#", style="dim", width=3)
        t.add_column("URL", max_width=60, overflow="ellipsis")
        t.add_column("Reason", max_width=20, style="yellow")
        icons = {"paywall": "💰", "cloudflare": "☁️", "captcha": "🤖",
                 "403": "🚫", "429_exhausted": "⏳", "timeout": "⏰",
                 "bot_detect": "🛡", "review_paper": "📖"}
        for i, bs in enumerate(self.state.blocked_sources, 1):
            url    = bs.get("url", bs.url if hasattr(bs, "url") else "")
            reason = bs.get("reason", bs.reason if hasattr(bs, "reason") else "")
            icon   = icons.get(reason, "❓")
            t.add_row(str(i), url[:60], f"{icon} {reason}")
        console.print(t)

    def _show_stats(self):
        """Show statistics about the extracted dataset."""
        if not self.state or not self.state.accepted_rows:
            console.print("  [yellow]No data yet. Run the pipeline first.[/yellow]")
            return
        rows = self.state.accepted_rows
        n    = len(rows)
        cols = [c for c in self.columns if c not in ("Source_URL", "row_confidence")]

        # Null rates per column
        null_rates = {
            col: sum(1 for r in rows
                     if not r.get(col) or str(r.get(col, "")).strip().lower()
                     in {"n/a", "none", "", "not specified", "unknown"})
                 / max(n, 1)
            for col in cols
        }

        t = Table(
            title=f"[bold green]📊 Dataset Stats — {n} rows[/bold green]",
            show_lines=False, border_style="green", box=box.SIMPLE,
        )
        t.add_column("Column", style="cyan")
        t.add_column("Fill %", style="white", width=8)
        t.add_column("Bar", style="white")

        for col, null_rate in sorted(null_rates.items(), key=lambda x: x[1], reverse=True):
            fill = 1.0 - null_rate
            bar  = "█" * int(fill * 20) + "░" * (20 - int(fill * 20))
            color = "green" if null_rate < 0.2 else "yellow" if null_rate < 0.5 else "red"
            t.add_row(col, f"[{color}]{fill:.0%}[/{color}]", f"[{color}]{bar}[/{color}]")

        console.print(t)
        console.print(f"\n  [dim]Sources: {len(self.state.documents)} | "
                      f"Blocked: {len(self.state.blocked_sources)} | "
                      f"Overall null: {sum(null_rates.values())/max(len(null_rates),1):.0%}[/dim]")

    def _export(self):
        if not self.state or not self.state.accepted_rows:
            console.print("  [yellow]No data to export yet.[/yellow]")
            return
        from tools.export import save_csv, save_json
        all_cols = self.columns + ["Source_URL", "row_confidence"]
        save_csv(self.state.accepted_rows, all_cols, self.state.output_path)
        save_json(self.state.accepted_rows, self.state.output_path)

    def _export_excel(self):
        """Export to Excel if openpyxl is available."""
        if not self.state or not self.state.accepted_rows:
            console.print("  [yellow]No data to export.[/yellow]")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            xlsx_path = Path(self.state.output_path).with_suffix(".xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dataset"
            all_cols = self.columns + ["Source_URL", "row_confidence"]
            fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
            font = Font(bold=True, color="FFFFFF")
            for ci, col in enumerate(all_cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.fill = fill
                cell.font = font
            for ri, row in enumerate(self.state.accepted_rows, 2):
                for ci, col in enumerate(all_cols, 1):
                    ws.cell(row=ri, column=ci, value=row.get(col, ""))
            wb.save(str(xlsx_path))
            console.print(f"  [green]✓ Excel saved: {xlsx_path}[/green]")
        except ImportError:
            console.print("  [yellow]pip install openpyxl to enable Excel export[/yellow]")

    def _smart_respond(self, message: str):
        context = self.memory.get_context(5)
        system = (
            "You are a helpful dataset building assistant chatbot.\n"
            f"Topic: {self.topic or 'not set'}\n"
            f"Columns: {self.columns or 'not set'}\n"
            f"Status: {'running' if self.running else 'idle'}\n"
            f"Domain context: {self.domain_context or 'none'}\n"
            f"Recent conversation:\n{context}\n\n"
            "Respond naturally and helpfully. Be concise."
        )
        response = self.client.complete(message, system=system, max_tokens=256)
        if response:
            console.print(f"\n  [bright_white]🤖 {response}[/bright_white]")
            self.memory.add("assistant", response)
        else:
            console.print("\n  [dim]I'm not sure how to help with that. Try /help.[/dim]")

    def _reset(self):
        if Confirm.ask("  [bold yellow]Reset current session?[/bold yellow]", default=False):
            self.topic = ""
            self.domain_context = ""
            self.columns = []
            self.state = None
            self.running = False
            console.print("  [green]✓ Session reset.[/green]")
            self._greet()

    def _goodbye(self):
        self.memory.add("system", "Session ended")
        self.memory.save()
        console.print("\n[bold bright_cyan]👋 Goodbye! Memory saved for next time.[/bold bright_cyan]\n")


def run_chatbot():
    bot = DatasetChatbot()
    bot.start()


if __name__ == "__main__":
    run_chatbot()
