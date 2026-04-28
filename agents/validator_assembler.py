"""
agents/validator_assembler.py — Agent 6: Quality Validation + Final Assembly.

Two-type validation:
  TYPE 1 — Structural/Format (rules-based, fast):
    Tier 1: Schema compliance (all columns present, no stray keys)
    Tier 2: Format checks (year, DOI, numeric fields, generic placeholders,
            DAR range, linker keywords)
    Tier 3: Cross-field consistency (DOI vs Source_URL)
    → Confidence re-scoring

  TYPE 2 — Semantic/Relevance (LLM-based, runs after Type 1):
    Given the user's topic + domain context, the LLM checks whether each
    accepted row is actually relevant and internally consistent.
    Rows flagged by LLM as "not relevant" are moved to rejected.
    This catches: wrong-topic extraction, hallucinated rows, copy-paste errors.

Source URL guarantee:
  Source_URL in output always comes from the pipeline's ingestion record,
  never from anything the LLM wrote.
"""

from __future__ import annotations

import json
import os
import re
import time
from collections import defaultdict
from pathlib import Path

from tools.console_setup import console
from agents.base_agent import BaseAgent
import config as cfg
from tools.llm_client import is_empty
from tools.export import (
    save_csv, save_rejected, save_json,
    display_preview, display_quality_report,
    display_rejected_rows,
)


DOI_REGEX = re.compile(r"10\.\d{4,}/[^\s,;\"'<>]+")

_LINKER_KEYWORDS = {
    "cleavable", "non-cleavable", "noncleavable", "val-cit", "valine",
    "hydrazone", "disulfide", "mcc", "smcc", "spdb", "maleimide",
    "pabc", "mc-", "peptide", "reducible", "acid-labile",
}
_GENERIC_VALS = {
    "not specified", "not mentioned", "various", "multiple",
    "see text", "refer to", "unknown", "not available",
    "not applicable", "n/a", "na", "none", "not provided",
}


class ValidatorAssemblerAgent(BaseAgent):

    def __init__(self, config: dict):
        super().__init__(agent_id="validator_assembler", config=config)
        self.output_path: str = str(config.get("output_path", "data/outputs/output.csv"))
        self._domain_context: str = str(config.get("domain_context", ""))

    async def run(self) -> dict | None:
        console.print("\n[bold cyan]✅ Agent 6: Validator + Assembler[/bold cyan] — starting")

        rows = self.read_input_queue("extracted_rows.jsonl")
        if not rows:
            console.print("[red]No data to validate.[/red]")
            self.log_status("run", "skip", {"reason": "no extracted rows"})
            return None

        columns = next((r.get("columns", []) for r in rows if r.get("columns")), [])
        columns = [c for c in columns if c not in ("Source_URL", "Dataset_Links")]
        topic   = next((r.get("topic", "") for r in rows if r.get("topic")), "")

        console.print(f"  [dim]{len(rows)} rows | {len(columns)} columns[/dim]")
        self.log_status("started", "success", {"rows": len(rows)})

        issues: dict[str, int] = defaultdict(int)

        # ── Type 1 — Structural Validation ───────────────────────────────

        # Tier 1: Schema compliance
        for row in rows:
            data = row.setdefault("data", {})
            for col in columns:
                if col not in data:
                    data[col] = "N/A"
                    issues["missing_key"] += 1
            # Strip LLM-hallucinated URL fields from data
            for bad_key in ["Source_URL", "source_url", "Source URL",
                            "URL", "url", "link", "Link"]:
                data.pop(bad_key, None)
            stray = [k for k in list(data.keys()) if k not in columns]
            for k in stray:
                del data[k]

        # Tier 2: Format checks
        for row in rows:
            data       = row.get("data", {})
            row_issues = row.setdefault("issues", [])
            for col in columns:
                val = data.get(col, "")
                if is_empty(val):
                    continue
                cl = col.lower()

                if any(kw in cl for kw in ["year", "date"]):
                    if not re.findall(r"\b(19|20)\d{2}\b", val):
                        row_issues.append(f"format:{col}:invalid_year")
                        issues["invalid_year"] += 1

                if "doi" in cl:
                    if not DOI_REGEX.match(val):
                        row_issues.append(f"format:{col}:invalid_doi")
                        issues["invalid_doi"] += 1

                if any(kw in cl for kw in [
                    "dose", "concentration", "quantity", "amount",
                    "ic50", "ec50", "ki", "kd", "score", "rating",
                    "p-value", "pvalue",
                ]):
                    if not re.search(r"\d", val):
                        row_issues.append(f"format:{col}:no_numeric")
                        issues["no_numeric"] += 1

                # ADC-specific: DAR must be 0.5–12
                if cl in ("dar", "drug-to-antibody ratio", "drug to antibody ratio"):
                    try:
                        dar_val = float(re.search(r"[\d.]+", val).group())
                        if not (0.5 <= dar_val <= 12.0):
                            row_issues.append(f"format:{col}:dar_out_of_range")
                            issues["dar_out_of_range"] += 1
                    except (AttributeError, ValueError):
                        row_issues.append(f"format:{col}:dar_not_numeric")
                        issues["dar_not_numeric"] += 1

                if "linker" in cl:
                    if not any(kw in val.lower() for kw in _LINKER_KEYWORDS):
                        row_issues.append(f"format:{col}:unrecognised_linker")
                        issues["unrecognised_linker"] += 1

                if any(kw in cl for kw in ["name", "title", "drug", "compound"]):
                    if val.strip().lower() in _GENERIC_VALS or len(val.strip()) < 2:
                        row_issues.append(f"format:{col}:generic_value")
                        issues["generic_value"] += 1

        # Tier 3: Cross-field consistency
        for row in rows:
            data       = row.get("data", {})
            row_issues = row.get("issues", [])
            source_url = row.get("source_url", "")   # from pipeline record
            doi_col    = next((c for c in columns if "doi" in c.lower()), None)
            if doi_col and source_url:
                doi_val   = data.get(doi_col, "")
                url_doi_m = DOI_REGEX.search(source_url)
                if doi_val and not is_empty(doi_val) and url_doi_m:
                    url_doi = url_doi_m.group()
                    if doi_val != url_doi and not url_doi.startswith(doi_val):
                        row_issues.append("cross_field:doi_mismatch")
                        issues["doi_mismatch"] += 1

        # Confidence re-scoring
        n_cols = max(len(columns), 1)
        for row in rows:
            score      = row.get("row_confidence", 1.0)
            confidence = row.get("confidence", {})
            data       = row.get("data", {})

            for issue in row.get("issues", []):
                if issue.startswith("format:"):
                    score -= 0.12
                elif issue.startswith("cross_field:"):
                    score -= 0.20
                elif "generic_value" in issue:
                    score -= 0.08
                elif "dar_out_of_range" in issue:
                    score -= 0.10

            high_conf  = sum(1 for v in confidence.values() if v == "high")
            missing    = sum(1 for v in confidence.values() if v == "missing")
            null_count = sum(1 for c in columns if is_empty(data.get(c, "N/A")))
            fill_rate  = 1.0 - (null_count / n_cols)
            fill_bonus = (fill_rate - 0.5) * 0.4

            conf_bonus = (high_conf * 0.02) - (missing * 0.08) + fill_bonus
            row["row_confidence"] = max(0.0, min(1.0, score + conf_bonus))

        # Global dedup
        rows = self._global_dedup(rows, columns)

        # ── Type 2 — LLM Semantic/Relevance Validation ───────────────────
        console.print(f"  [dim]Type 2: LLM relevance check for {topic}…[/dim]")
        rows = self._llm_relevance_check(rows, columns, topic)

        # Split accepted / rejected
        accepted: list[dict] = []
        rejected: list[dict] = []
        for row in rows:
            flat = self._to_flat(row, columns)
            if row.get("row_confidence", 0.0) >= 0.3:
                accepted.append(flat)
            else:
                rejected.append(flat)

        # Add blocked sources to rejected log
        blocked_records = self.read_input_queue("triage_blocked.jsonl")
        for bs in blocked_records:
            rejected.append({
                col: "N/A" for col in columns
            } | {
                "Source_URL":     bs.get("url", ""),
                "row_confidence": 0.0,
                "_issues":        f"blocked:{bs.get('reason', 'unknown')}",
                "_block_reason":  bs.get("reason", ""),
            })

        # Write state bus
        validated_path = self.state_dir / "validated_rows.jsonl"
        validated_path.write_text("", encoding="utf-8")
        for flat in accepted:
            self.write_output("validated_rows.jsonl", flat)

        rejected_path = self.state_dir / "rejected_rows.jsonl"
        rejected_path.write_text("", encoding="utf-8")
        for flat in rejected:
            self.write_output("rejected_rows.jsonl", flat)

        null_rates: dict[str, float] = {
            col: sum(1 for r in accepted if is_empty(r.get(col, "N/A")))
                 / max(len(accepted), 1)
            for col in columns
        }
        overall_null       = sum(null_rates.values()) / max(len(null_rates), 1)
        needs_more_sources = overall_null > cfg.MAX_NULL_RATE

        # Write final CSV + JSON
        all_cols = columns + ["Source_URL", "row_confidence"]
        save_csv(accepted, all_cols, self.output_path)
        save_json(accepted, self.output_path)
        self._try_save_excel(accepted, all_cols)

        if rejected:
            save_rejected(rejected, os.path.dirname(self.output_path) or str(cfg.OUTPUT_DIR))

        display_preview(accepted, all_cols)
        display_quality_report(
            accepted=len(accepted), rejected=len(rejected),
            null_rates=null_rates, issues=dict(issues),
        )
        if rejected:
            display_rejected_rows(rejected, all_cols)

        self.log_status("completed", "success", {
            "accepted":    len(accepted), "rejected":   len(rejected),
            "blocked":     len(blocked_records),
            "null_rate":   f"{overall_null:.0%}",
            "output_path": self.output_path,
        })

        if needs_more_sources:
            console.print(
                f"  [yellow]⚠ High null rate ({overall_null:.0%}) "
                "— orchestrator should trigger more sources[/yellow]"
            )

        console.print(
            f"\n[bold cyan]✅ Validator Assembler[/bold cyan] — done. "
            f"{len(accepted)} accepted, {len(rejected)} rejected."
        )

        return {
            "null_rate":          overall_null,
            "needs_more_sources": needs_more_sources,
            "null_rates":         null_rates,
            "accepted":           len(accepted),
            "rejected":           len(rejected),
        }

    # ── Type 2: LLM Relevance Check ──────────────────────────────────────

    def _llm_relevance_check(self, rows: list[dict], columns: list[str],
                               topic: str) -> list[dict]:
        """
        LLM checks each row for relevance to the topic.
        Rows that are clearly irrelevant or hallucinated get confidence penalised.
        Runs in batches of 10 to limit API calls.
        """
        if not topic or not rows:
            return rows

        dom_ctx = self._domain_context
        system = (
            f'You are a scientific data quality auditor for the topic: "{topic}".\n'
            + (f'Domain context: {dom_ctx}\n' if dom_ctx else "")
            + "You will receive a JSON array of extracted rows.\n"
            "For each row, decide if it is RELEVANT to the topic.\n"
            "Mark as NOT_RELEVANT if:\n"
            "  - The row is clearly from the wrong scientific field\n"
            "  - The identity fields (name, compound, etc.) don't match the topic at all\n"
            "  - Most values look like they were hallucinated (too generic or inconsistent)\n"
            "Return a JSON array of {index, relevant: true/false, reason: str}."
        )

        BATCH = 10
        for i in range(0, len(rows), BATCH):
            batch = rows[i:i+BATCH]
            batch_data = [
                {"index": i+j, "data": {col: row.get("data", {}).get(col, "N/A")
                                        for col in columns[:8]}}
                for j, row in enumerate(batch)
            ]
            try:
                result = self.llm.complete_json(
                    json.dumps(batch_data, indent=2),
                    system=system,
                    max_tokens=1024,
                )
                if isinstance(result, list):
                    for item in result:
                        if not isinstance(item, dict):
                            continue
                        idx     = item.get("index", -1) - i
                        relevant = item.get("relevant", True)
                        reason  = item.get("reason", "")
                        if 0 <= idx < len(batch) and not relevant:
                            batch[idx]["row_confidence"] = min(
                                batch[idx].get("row_confidence", 0.5),
                                0.25,
                            )
                            batch[idx]["issues"].append(
                                f"llm_relevance:not_relevant:{reason[:60]}"
                            )
                            console.print(
                                f"  [dim yellow]⚠ Row flagged as not relevant: "
                                f"{reason[:80]}[/dim yellow]"
                            )
            except Exception as e:
                console.print(f"  [dim yellow]LLM relevance check skipped: {e}[/dim yellow]")

        return rows

    # ── Helpers ───────────────────────────────────────────────────────────

    def _global_dedup(self, rows: list[dict], columns: list[str]) -> list[dict]:
        seen:   set[str] = set()
        deduped: list    = []
        for row in rows:
            data = row.get("data", {})
            sig  = "|".join(str(data.get(c, "")).lower().strip() for c in columns)
            if sig not in seen:
                seen.add(sig)
                deduped.append(row)
        if len(deduped) < len(rows):
            console.print(f"  [dim]Deduped: {len(rows)} → {len(deduped)} rows[/dim]")
        return deduped

    def _to_flat(self, row: dict, columns: list[str]) -> dict:
        """Source URL comes ONLY from pipeline record, never from LLM output."""
        flat = dict(row.get("data", {}))
        # Remove any LLM-generated URL fields
        flat.pop("Source_URL", None)
        flat.pop("source_url", None)
        flat.pop("Source URL", None)
        # Set authoritative Source_URL from ingestion record
        flat["Source_URL"]     = row.get("source_url", "")
        flat["row_confidence"] = round(row.get("row_confidence", 0.0), 2)
        if row.get("issues"):
            flat["_issues"] = "; ".join(row["issues"])
        return flat

    def _try_save_excel(self, rows: list[dict], columns: list[str]) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            xlsx_path = Path(self.output_path).with_suffix(".xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Data"
            fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
            font = Font(bold=True, color="FFFFFF")
            for ci, col in enumerate(columns, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(rows, 2):
                for ci, col in enumerate(columns, 1):
                    ws.cell(row=ri, column=ci, value=row.get(col, ""))
            for ci in range(1, len(columns) + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 18
            wb.save(str(xlsx_path))
            console.print(f"  [dim]📊 Excel saved: {xlsx_path.name}[/dim]")
        except ImportError:
            pass
        except Exception as e:
            console.print(f"  [dim yellow]Excel export skipped: {e}[/dim yellow]")
